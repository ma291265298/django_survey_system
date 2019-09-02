from django.forms import model_to_dict
from django.http import JsonResponse
from django.shortcuts import render
import json
import base64
import smtplib
import random
import string
import openpyxl
from email.mime.text import MIMEText
from email.utils import formataddr

from Crypto.Cipher import PKCS1_v1_5 as Cipher_pkcs1_v1_5
from Crypto.PublicKey import RSA
from survey.models import *
from django.db.models import Q

import time
import threading

import traceback

my_sender = '5724924@qq.com'  # 发件人邮箱账号
my_pass = 'rtccbnjnydebbigg'  # 发件人邮箱密码

exitFlag=0
class checkSensitiveThread (threading.Thread):
    def __init__(self, paper_obj):
        threading.Thread.__init__(self)
        self.paper_obj=paper_obj
    def run(self):
        print ("开始线程")
        checkSensitiveAI(self.paper_obj)
        print ("退出线程")

def checkSensitiveAI(paper_obj):
    time.sleep(3)
    message=checkSensitiveFunction(paper_obj)
    if message is True:
        paper_obj.verify='未发布'
    else:
        paper_obj.verify = '审核未通过,原因:'+message[0]+",识别到:"+message[1]
    paper_obj.save()

def checkSensitiveFunction(paper_obj):
    a = getModifyQuestion({'paperId': paper_obj.id})
    check = []
    check.append(a['paperName'])
    check.append(a['paperDetail'])
    for i in a['questions']:
        check.append(i['questionName'])
        for k in i['options']:
            check.append(k)
    wb = openpyxl.load_workbook("sensitive.xlsx")
    sheet = wb.active
    for i in range(sheet.max_row):
        for k in check:
            if str(sheet.cell(i + 1, 4).value) in str(k):
                return [str(sheet.cell(i + 1, 2).value),str(sheet.cell(i + 1, 4).value)]
    return True

def loginCheck(func):
    def wrapper(request, *args, **kwargs):
        user = request.session.get("userID")
        if not user:
            return render(request, "login.html")
        else:
            return func(request, *args, **kwargs)

    return wrapper


@loginCheck
def addView(request):
    return render(request, "add.html")


def erro(request):
    return render(request, "404.html")


def index(request):
    return render(request, "login.html")


def addSuccess(request):
    if not request.session.get("userID"):
        return JsonResponse({'resultCode': 1})
    user = User.objects.filter(id=request.session.get("userID"))
    paper_obj = Paper(name=request.POST.get('paperName'),
                      detail=request.POST.get('paperDetail'), uid=user[0], verify='AI审核中')
    paper_obj.save()
    for i in json.loads(request.POST.get('questions')):
        question_obj = Question(no=i['no'], content=i['questionName'], type=i['type'], ismustfill=i['ismustfill'],
                                pid=paper_obj)
        question_obj.save()
        count = 1
        for k in i['option']:
            option_obj = Option(no=count, content=k, qid=question_obj)
            count += 1
            option_obj.save()
    # 创建新线程
    thread1 = checkSensitiveThread(paper_obj)
    # 开启新线程
    thread1.start()
    print("退出主线程")
    return JsonResponse({'resultCode': 0})


def sendRegisterEmail(request):
    f = open('first_private_key.txt', 'r', encoding='utf-8')
    firstPrivateKey = f.read()
    f.close()
    f = open('second_public_key.txt', 'r', encoding='utf-8')
    secondPublicKey = f.read()
    f.close()
    m = request.POST.get('raw')
    adict = rsaDecrypt(m, firstPrivateKey)
    adict['whoisyourdaddy'] = 'sb110'
    url = rsaCrypto(adict, secondPublicKey)
    mail(adict, url)


def secondRegister(request, m):
    f = open('second_private_key.txt', 'r', encoding='utf-8')
    secondPrivateKey = f.read()
    f.close()
    print("注册信息")
    try:
        adict = rsaDecrypt(m, secondPrivateKey)
        if adict['whoisyourdaddy'] == 'sb110':
            user_obj = User.objects.filter(Q(name=adict['name']) | Q(email=adict['email']))
            if len(user_obj) == 0:
                user_obj = User(name=adict['name'], email=adict['email'], password=adict['password'])
                print(adict)
                user_obj.save()
                return render(request, "registerSuccess.html")
            return render(request, "registerDumplicate.html")
    except:
        pass
    return render(request, "404.html")


def mail(adict, url):
    ret = True
    try:
        content = '\
                <h2>欢迎注册sb110问卷网, 离注册完成还有最后一步</h2>\
                <p><a href="http://47.100.167.60/a/AkdjrEkclaoq/' + url + '">点击此处以激活账号</a></p>\
                <h2>注意,此链接仅可使用一次</h2>\
                <h2>如果这并不是你本人操作, 请忽略这封邮件</h2>'
        msg = MIMEText(content, 'html', 'utf-8')
        msg['From'] = "SB110Rigiseter <5724924@qq.com>"  # 括号里的对应发件人邮箱昵称、发件人邮箱账号
        msg['To'] = formataddr(["FK", adict['email']])  # 括号里的对应收件人邮箱昵称、收件人邮箱账号
        msg['Subject'] = "激活你的sb110账号"  # 邮件的主题，也可以说是标题
        server = smtplib.SMTP_SSL("smtp.qq.com", 465)  # 发件人邮箱中的SMTP服务器，端口是25
        server.login(my_sender, my_pass)  # 括号中对应的是发件人邮箱账号、邮箱密码
        server.sendmail(my_sender, [adict['email'], ], msg.as_string())  # 括号中对应的是发件人邮箱账号、收件人邮箱账号、发送邮件
        server.quit()  # 关闭连接
    except Exception:  # 如果 try 中的语句没有执行，则会执行下面的 ret=False
        ret = False
    return ret


def rsaDecrypt(m, rsa_private_key):
    rsakey = RSA.importKey(rsa_private_key.encode())
    cipher = Cipher_pkcs1_v1_5.new(rsakey)
    text = cipher.decrypt(base64.b64decode(m), None)
    c = text.decode('utf8')
    return eval(c)


def rsaCrypto(adict, rsa_public_key):
    rsakey = RSA.importKey(rsa_public_key.encode())
    cipher = Cipher_pkcs1_v1_5.new(rsakey)
    cipher_text = base64.b64encode(cipher.encrypt(str(adict).encode()))
    return cipher_text.decode('utf8')


def nameCheck(request):
    if len(User.objects.filter(name=request.POST.get('name'))) == 0:
        return JsonResponse({'resultCode': 0})
    return JsonResponse({'resultCode': 1})


def emailCheck(request):
    if len(User.objects.filter(email=request.POST.get('email'))) == 0:
        return JsonResponse({'resultCode': 0})
    return JsonResponse({'resultCode': 1})


def loginAction(request):
    user = User.objects.filter(Q(name=request.POST.get('name')) | Q(email=request.POST.get('name')),
                               password=request.POST.get('password'))
    if len(user) == 0:
        return JsonResponse({'resultCode': 1})
    request.session['userID'] = user[0].id
    print(user[0])
    return JsonResponse({'resultCode': 0})


@loginCheck
def userView(request):
    Paperobj = Paper.objects.filter(uid=request.session.get("userID"))
    rsa = []
    f = open('second_public_key.txt', 'r', encoding='utf-8')
    secondPublicKey = f.read()
    f.close()
    for i in list(Paperobj):
        adict = {"userId": request.session.get("userID"), "paperId": i.id, "whoisyourdaddy": "sb110"}  # 发布
        bdict = {"userId": request.session.get("userID"), "paperId": i.id, "whoisyourdaddy": "sb111"}  # 编辑
        cdict = {"userId": request.session.get("userID"), "paperId": i.id, "whoisyourdaddy": "sb112"}  # 分享
        ddict = {"userId": request.session.get("userID"), "paperId": i.id, "whoisyourdaddy": "sb113"}  # 统计
        edict = {"userId": request.session.get("userID"), "paperId": i.id, "whoisyourdaddy": "sb114"}  # 删除
        fdict = {"userId": request.session.get("userID"), "paperId": i.id, "whoisyourdaddy": "sb115"}  # 停止
        rsa.append((i, rsaCrypto(adict, secondPublicKey),
                    rsaCrypto(bdict, secondPublicKey),
                    rsaCrypto(cdict, secondPublicKey),
                    rsaCrypto(ddict, secondPublicKey),
                    rsaCrypto(edict, secondPublicKey),
                    getAnswerPeople(i.id),
                    rsaCrypto(fdict, secondPublicKey)
                    ))
    return render(request, "user.html", {"papers": rsa})


def getAnswerPeople(id):
    Answer_obj = Answer.objects.filter(pid__id=id)
    if len(Answer_obj) != 0:
        return Answer_obj.filter(qid=Answer_obj[0].qid).count()
    return 0


@loginCheck
def modifyView(request, m):
    f = open('second_private_key.txt', 'r', encoding='utf-8')
    secondPrivateKey = f.read()
    f.close()
    try:
        adict = rsaDecrypt(m, secondPrivateKey)
        if not adict['whoisyourdaddy'] == 'sb111':
            return render(request, "404.html")
        if adict['userId'] == request.session.get("userID"):
            return render(request, "modify.html", getModifyQuestion(adict))
        return render(request, "404.html")
    except Exception as err:
        print(err)
        return render(request, "404.html")


def getModifyQuestion(adict):
    questionInfo = []
    paper = Paper.objects.filter(id=adict['paperId'])
    questions = Question.objects.filter(pid=paper[0])
    for i in questions:
        questionDict = {'no': i.no,
                        'type': i.type,
                        'questionName': i.content,
                        'options': [],
                        'ismustfill': i.ismustfill, }
        options = Option.objects.filter(qid=i)
        for k in options:
            questionDict['options'].append(k.content)
        questionInfo.append(questionDict)
    return {"paperId": paper[0].id, "paperName": paper[0].name, "paperDetail": paper[0].detail,
            "questions": questionInfo}


def deleteAction(request):
    f = open('second_private_key.txt', 'r', encoding='utf-8')
    secondPrivateKey = f.read()
    f.close()
    try:
        m = request.POST.get('raw')
        adict = rsaDecrypt(m, secondPrivateKey)
        if not adict['whoisyourdaddy'] == 'sb114':
            return JsonResponse({'resultCode': 1})
        if adict['userId'] == request.session.get("userID"):
            deleteFunction(adict['paperId'])
            return JsonResponse({'resultCode': 0})
        return render(request, "404.html")
    except Exception as err:
        print(err)
        JsonResponse({'resultCode': 1})

def stopAction(request):
    f = open('second_private_key.txt', 'r', encoding='utf-8')
    secondPrivateKey = f.read()
    f.close()
    try:
        m = request.POST.get('raw')
        adict = rsaDecrypt(m, secondPrivateKey)
        if not adict['whoisyourdaddy'] == 'sb115':
            return JsonResponse({'resultCode': 1})
        if adict['userId'] == request.session.get("userID"):
            stopFunction(adict['paperId'])
            return JsonResponse({'resultCode': 0})
        return render(request, "404.html")
    except Exception as err:
        print(err)
        JsonResponse({'resultCode': 1})


def deleteFunction(paperId):
    Paper_obj = Paper.objects.get(id=paperId)
    Paper_obj.delete()

def stopFunction(paperId):
    Paper_obj = Paper.objects.get(id=paperId)
    Paper_obj.verify='已停止'
    Paper_obj.save()



def saveModifyAction(request):
    deleteFunction(request.POST.get('paperId'))
    return addSuccess(request)


def releasePaperAction(request):
    f = open('second_private_key.txt', 'r', encoding='utf-8')
    secondPrivateKey = f.read()
    f.close()
    try:
        m = request.POST.get('raw')
        adict = rsaDecrypt(m, secondPrivateKey)
        if not adict['whoisyourdaddy'] == 'sb110':
            return JsonResponse({'resultCode': 1})
        if adict['userId'] == request.session.get("userID"):
            releasePaperFunction(adict['paperId'])
            return JsonResponse({'resultCode': 0})
        return JsonResponse({'resultCode': 1})
    except Exception as err:
        print(err)
        return JsonResponse({'resultCode': 1})


def releasePaperFunction(paperId):
    Paper_obj = Paper.objects.get(id=paperId)
    Paper_obj.verify = '已发布'
    Paper_obj.url = urlGenerator()
    Paper_obj.save()


def urlGenerator(size=8, chars=string.ascii_letters + string.digits):
    return ''.join(random.choice(chars) for _ in range(size))


def paperView(request, m):
    try:
        paper_obj = Paper.objects.get(url=m,verify='已发布')
    except:
        return render(request, "404.html")
    return render(request, "paper.html", getModifyQuestion({"paperId": paper_obj.id}))


def answer(request):
    try:
        paper_obj = Paper.objects.get(id=request.POST.get('paperId'))
        questions_obj = Question.objects.filter(pid=paper_obj)
        questions_obj
        for i, j in zip(json.loads(request.POST.get('answer')), questions_obj):
            for k in i:
                answer_obj = Answer(content=k, qid=j, pid=paper_obj)
                answer_obj.save()
    except:
        return JsonResponse({'resultCode': 1})

    return JsonResponse({'resultCode': 0})


def anserSuccessView(request):
    return render(request, "success.html")


@loginCheck
def summaryView(request, m):
    f = open('second_private_key.txt', 'r', encoding='utf-8')
    secondPrivateKey = f.read()
    f.close()
    try:
        adict = rsaDecrypt(m, secondPrivateKey)
        if not adict['whoisyourdaddy'] == 'sb113':
            return render(request, "404.html")
        if adict['userId'] == request.session.get("userID"):
            return render(request, "summary.html", getSummaryFunction(adict['paperId']))
        return render(request, "404.html")
    except Exception as err:
        print(err)
        return render(request, "404.html")


def getSummaryFunction(id):
    summary = []
    question_obj = Question.objects.filter(pid__id=id)
    for i in question_obj:
        adict = {}
        adict['st'] = urlGenerator()
        adict['questionName'] = i.content
        adict['type'] = i.type
        if i.type == '单选' or i.type == '多选':
            adict['option'] = json.dumps([j.content for j in Option.objects.filter(qid=i)], ensure_ascii=False)
            adict['number'] = json.dumps(
                [Answer.objects.filter(content=j.content).filter(qid=i).count() for j in Option.objects.filter(qid=i)],
                ensure_ascii=False)
        if i.type == '自由':
            adict['content'] = json.dumps([j.content for j in Answer.objects.filter(qid=i)], ensure_ascii=False)
            adict['length'] = len(adict['content'])
            if adict['content'] == '[]' or adict['content'] == '[null]':
                adict['length'] = 0
        summary.append(adict)
    return {"answer": summary}



